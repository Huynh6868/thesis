"""
CPLEX Results to Excel Exporter
Combines Phase 1 (team assignments) and Phase 2 (room assignments) results
into formatted Excel file.

Usage:
    python export_results_to_excel.py
"""

import re
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==================== CONFIGURATION ====================
SCRIPT_DIR = Path(__file__).parent

# Check if we're in auto_result directory
if SCRIPT_DIR.name == "auto_result":
    AUTO_RESULT_DIR = SCRIPT_DIR
else:
    # Try to find auto_result from initial solution or other locations
    AUTO_RESULT_DIR = SCRIPT_DIR.parent / "model1_revise1" / "auto_result"
    if not AUTO_RESULT_DIR.exists():
        AUTO_RESULT_DIR = SCRIPT_DIR / "auto_result"

PHASE1_RESULTS = AUTO_RESULT_DIR / "phase1_results.dat"
PHASE2_RESULTS = AUTO_RESULT_DIR / "phase2_results.dat"
DATA_FILE = AUTO_RESULT_DIR / "Data_1_EDITED.dat"
OUTPUT_EXCEL = AUTO_RESULT_DIR / "surgery_schedule.xlsx"

print(f"Working directory: {AUTO_RESULT_DIR}")

# Surgery type mapping (1-10)
SURGERY_TYPE_MAP = {
    1: "adenotonsillectomy",
    2: "microlaryngoscopy",
    3: "buccal mucosa bioppsy",
    4: "excision of the lymphadenopathy from the lumbar",
    5: "septoplasty",
    6: "modified radical mastoidectomy",
    7: "thyroidectomy",
    8: "rhinoplasty",
    9: "endoscopic sinus",
    10: "sleep apnea diagnosis test"
}

# ==================== PARSER FUNCTIONS ====================

def parse_matrix_3d(file_path, matrix_name):
    """Parse 3D matrix [s][p][d] from .dat file."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    pattern = rf'{matrix_name}\s*=\s*\[(.*?)\];'
    match = re.search(pattern, content, re.DOTALL)
    if not match:
        raise ValueError(f"Matrix '{matrix_name}' not found in {file_path}")
    
    matrix_str = match.group(1)
    matrix = {}
    
    # Remove all whitespace and newlines for easier parsing
    matrix_str = ''.join(matrix_str.split())
    
    # Find all surgeon blocks: each surgeon is wrapped in [...]
    # We need to match surgeon blocks at the top level
    s_idx = 0
    depth = 0
    start = -1
    
    for i, char in enumerate(matrix_str):
        if char == '[':
            if depth == 0:
                start = i
            depth += 1
        elif char == ']':
            depth -= 1
            if depth == 0 and start != -1:
                # Extract surgeon block
                surgeon_block = matrix_str[start+1:i]
                
                # Now parse patient blocks within this surgeon
                p_idx = 0
                p_depth = 0
                p_start = -1
                
                for j, c in enumerate(surgeon_block):
                    if c == '[':
                        if p_depth == 0:
                            p_start = j
                        p_depth += 1
                    elif c == ']':
                        p_depth -= 1
                        if p_depth == 0 and p_start != -1:
                            # Extract patient block
                            patient_block = surgeon_block[p_start+1:j]
                            
                            # Parse day values
                            day_values = [x.strip() for x in patient_block.split(',') if x.strip()]
                            for d_idx, day_val in enumerate(day_values):
                                try:
                                    if matrix_name == 'startsp_in':
                                        value = float(day_val)
                                    else:
                                        value = int(day_val)
                                    matrix[(s_idx + 1, p_idx + 1, d_idx + 1)] = value
                                except ValueError:
                                    pass  # Skip invalid values
                            
                            p_idx += 1
                
                s_idx += 1
                start = -1
    
    return matrix


def parse_room_matrix(file_path):
    """Parse room assignment matrix v[p][k][d] from Phase 2 results."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    pattern = r'v_in\s*=\s*\[(.*?)\];'
    match = re.search(pattern, content, re.DOTALL)
    if not match:
        raise ValueError(f"Matrix 'v_in' not found in {file_path}")
    
    matrix_str = match.group(1)
    matrix = {}
    
    # Remove all whitespace and newlines for easier parsing
    matrix_str = ''.join(matrix_str.split())
    
    # Find all patient blocks at top level
    p_idx = 0
    depth = 0
    start = -1
    
    for i, char in enumerate(matrix_str):
        if char == '[':
            if depth == 0:
                start = i
            depth += 1
        elif char == ']':
            depth -= 1
            if depth == 0 and start != -1:
                # Extract patient block
                patient_block = matrix_str[start+1:i]
                
                # Now parse room blocks within this patient
                k_idx = 0
                k_depth = 0
                k_start = -1
                
                for j, c in enumerate(patient_block):
                    if c == '[':
                        if k_depth == 0:
                            k_start = j
                        k_depth += 1
                    elif c == ']':
                        k_depth -= 1
                        if k_depth == 0 and k_start != -1:
                            # Extract room block
                            room_block = patient_block[k_start+1:j]
                            
                            # Parse day values
                            day_values = [x.strip() for x in room_block.split(',') if x.strip()]
                            for d_idx, day_val in enumerate(day_values):
                                try:
                                    value = int(day_val)
                                    matrix[(p_idx + 1, k_idx + 1, d_idx + 1)] = value
                                except ValueError:
                                    pass
                            
                            k_idx += 1
                
                p_idx += 1
                start = -1
    
    return matrix


def parse_patient_types(data_file_path):
    """Parse PatientType array from data file."""
    with open(data_file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    pattern = r'PatientType\s*=\s*\[(.*?)\];'
    match = re.search(pattern, content, re.DOTALL)
    if not match:
        raise ValueError(f"PatientType not found in {data_file_path}")
    
    types_str = match.group(1)
    types = [int(x.strip()) for x in types_str.split(',') if x.strip()]
    
    return {i + 1: t for i, t in enumerate(types)}


def minutes_to_hhmm(minutes):
    """Convert minutes to HH:MM format.
    Adds 8-hour offset to start from 8:00 AM instead of 0:00.
    """
    # Add 8 hours (480 minutes) to match rule-based simulation time
    adjusted_minutes = minutes + 480
    h = int(adjusted_minutes // 60)
    m = int(adjusted_minutes % 60)
    return f"{h}:{m:02d}"


# ==================== MAIN CONVERSION ====================

def extract_schedule_data():
    """Extract and combine data from Phase 1 and Phase 2."""
    print(f"[1/5] Reading Phase 1 results from {PHASE1_RESULTS}...")
    wsp = parse_matrix_3d(PHASE1_RESULTS, 'wsp_in')
    ysp = parse_matrix_3d(PHASE1_RESULTS, 'ysp_in')
    zsp = parse_matrix_3d(PHASE1_RESULTS, 'zsp_in')
    startsp = parse_matrix_3d(PHASE1_RESULTS, 'startsp_in')
    
    print(f"[2/5] Reading Phase 2 results from {PHASE2_RESULTS}...")
    v_matrix = parse_room_matrix(PHASE2_RESULTS)
    
    print(f"[3/5] Reading patient types from {DATA_FILE}...")
    patient_types = parse_patient_types(DATA_FILE)
    
    print(f"[4/5] Building schedule...")
    
    num_patients = max(p for s, p, d in wsp.keys())
    num_surgeons = max(s for s, p, d in wsp.keys())
    num_days = max(d for s, p, d in wsp.keys())
    
    schedule = []
    
    for p in range(1, num_patients + 1):
        # Find which day this patient is scheduled
        scheduled_day = None
        main_surgeon = None
        assist1_surgeon = None
        assist2_surgeon = None
        start_time = 0
        room = None
        
        for d in range(1, num_days + 1):
            for s in range(1, num_surgeons + 1):
                if wsp.get((s, p, d), 0) == 1:
                    main_surgeon = s
                    scheduled_day = d
                    start_time = startsp.get((s, p, d), 0)
                
                if ysp.get((s, p, d), 0) == 1:
                    assist1_surgeon = s
                
                if zsp.get((s, p, d), 0) == 1:
                    assist2_surgeon = s
            
            # Find room
            if scheduled_day == d:
                for k in [1, 2]:
                    if v_matrix.get((p, k, d), 0) == 1:
                        room = k
                        break
        
        if scheduled_day is None:
            print(f"  WARNING: Patient P{p} not scheduled!")
            continue
        
        # Get surgery type
        surgery_type_code = patient_types[p]
        surgery_type_name = SURGERY_TYPE_MAP[surgery_type_code]
        
        schedule.append({
            'pid': f'P{p}',
            'surgery_type': surgery_type_name,
            'day': scheduled_day - 1,  # 0-indexed for display
            'time_hhmm': minutes_to_hhmm(start_time),
            'room': room if room else 1,
            'main': f'S{main_surgeon}',
            'assist1': f'S{assist1_surgeon}',
            'assist2': f'S{assist2_surgeon}'
        })
    
    # Sort by day, then time
    schedule.sort(key=lambda x: (x['day'], x['time_hhmm']))
    
    return schedule


def export_to_excel(schedule):
    """Export schedule to formatted Excel file."""
    print(f"[5/5] Exporting to Excel: {OUTPUT_EXCEL}...")
    
    # Create DataFrame
    df = pd.DataFrame(schedule)
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Surgery Schedule"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    headers = ['pid', 'surgery_type', 'day', 'time_hhmm', 'room', 'main', 'assist1', 'assist2']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, record in enumerate(schedule, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record[header])
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # pid
        'B': 45,  # surgery_type
        'C': 8,   # day
        'D': 12,  # time_hhmm
        'E': 8,   # room
        'F': 8,   # main
        'G': 10,  # assist1
        'H': 10   # assist2
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save
    wb.save(OUTPUT_EXCEL)
    print(f"\n[OK] Successfully exported {len(schedule)} surgeries to {OUTPUT_EXCEL}")


# ==================== MAIN ====================

def main():
    print("=" * 70)
    print("CPLEX Results to Excel Exporter")
    print("=" * 70)
    
    # Check if input files exist
    for file_path in [PHASE1_RESULTS, PHASE2_RESULTS, DATA_FILE]:
        if not file_path.exists():
            print(f"ERROR: Input file not found: {file_path}")
            print("Please run CPLEX optimization first!")
            return 1
    
    try:
        schedule = extract_schedule_data()
        
        # Debug: Show room distribution
        room_counts = {}
        for surg in schedule:
            room = surg['room']
            room_counts[room] = room_counts.get(room, 0) + 1
        
        print(f"\nRoom distribution:")
        for room, count in sorted(room_counts.items()):
            print(f"  Room {room}: {count} surgeries")
        
        export_to_excel(schedule)
        
        print("\nFirst 5 surgeries:")
        for surg in schedule[:5]:
            print(f"  {surg}")
        
        return 0
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
