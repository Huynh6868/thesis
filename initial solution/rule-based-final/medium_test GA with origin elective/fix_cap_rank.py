"""
Fix Cap_Rank.xlsx - Add S16 to Assistant 2 for Medium Scale

This script adds S16 to the assistant 2 capabilities for all surgery types.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('Cap_Rank.xlsx')

# Check available sheet names
print("Available sheets:", wb.sheetnames)

# Get the capabilities sheet for medium scale
if 'capabilities med' in wb.sheetnames:
    ws = wb['capabilities med']
    print("\nProcessing 'capabilities med' sheet...")
    
    # Read the sheet into a DataFrame to understand structure
    df = pd.read_excel('Cap_Rank.xlsx', sheet_name='capabilities med')
    print("\nCurrent sheet structure:")
    print(df.head())
    
    # Find the assistant 2 column
    print("\nColumn names:", df.columns.tolist())
    
    # Add S16 to assistant 2 for each surgery type
    for idx, row in df.iterrows():
        if 'assistant 2' in df.columns or 'Assistant 2' in df.columns:
            col_name = 'assistant 2' if 'assistant 2' in df.columns else 'Assistant 2'
            current_a2 = str(df.at[idx, col_name])
            
            # Add S16 if not already there
            if 'S16' not in current_a2:
                if pd.notna(current_a2) and current_a2 != 'nan':
                    # Append S16 to existing list
                    df.at[idx, col_name] = current_a2 + ', S16'
                    print(f"Row {idx}: Added S16 to '{current_a2}' -> '{df.at[idx, col_name]}'")
                else:
                    df.at[idx, col_name] = 'S16'
                    print(f"Row {idx}: Set assistant 2 to 'S16'")
    
    # Write back to Excel
    with pd.ExcelWriter('Cap_Rank.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='capabilities med', index=False)
    
    print("\n✓ Successfully updated 'capabilities med' sheet")
    print("✓ Added S16 to assistant 2 for all surgery types")
else:
    print("ERROR: 'capabilities med' sheet not found!")
    print("Available sheets:", wb.sheetnames)

wb.close()

# Verify the fix
print("\n" + "="*60)
print("VERIFICATION")
print("="*60)

import rule_based_or_sim_v3 as sim
cap = sim.load_cap_rank_xlsx('Cap_Rank.xlsx')
a2 = cap.a2_by_type.get('thyroidectomy', set())
print(f"Assistant 2 for thyroidectomy: {sorted(a2)}")
print(f"Has S16: {'S16' in a2}")
print(f"Total assistant 2 surgeons: {len(a2)}")

if 'S16' in a2:
    print("\n✅ SUCCESS: S16 has been added to assistant 2!")
else:
    print("\n❌ FAILED: S16 still not in assistant 2")
