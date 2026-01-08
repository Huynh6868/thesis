"""
Doctor Schedule Generator
Generates weekly work schedules for 12 doctors with on-call rotations
Adheres to constraints: compensation days, shift limits, coverage requirements
"""

import random
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Set
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuration
NUM_DOCTORS = 20
SHIFTS_PER_WEEK = 2  # Each doctor must have exactly 2 on-call shifts per week
MIN_PRIMARY_SURGEONS_PER_SHIFT = 2  # Minimum surgeons from main surgeon group (1-8, 13-15, 17-19) per shift

# Days of the week
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
WEEKDAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
WEEKEND = ['Saturday', 'Sunday']

# On-call shift types
SHIFT_TYPES = {
    'Monday_Evening': 0,
    'Tuesday_Evening': 1,
    'Wednesday_Evening': 2,
    'Thursday_Evening': 3,
    'Friday_Evening': 4,
    'Saturday_Full': 5,
    'Sunday_Full': 6
}

# Doctor definitions
class Doctor:
    def __init__(self, id: int, name: str, can_be_main_surgeon: bool):
        self.id = id
        self.name = name
        self.can_be_main_surgeon = can_be_main_surgeon
    
    def __repr__(self):
        return f"Doctor {self.id}: {self.name} (Main: {self.can_be_main_surgeon})"

# Initialize 20 doctors: 14 main surgeons (1-8, 13-15, 17-19) + 6 assistants (9-12, 16, 20)
DOCTORS = [
    Doctor(1, "Surgeon 1", True),
    Doctor(2, "Surgeon 2", True),
    Doctor(3, "Surgeon 3", True),
    Doctor(4, "Surgeon 4", True),
    Doctor(5, "Surgeon 5", True),
    Doctor(6, "Surgeon 6", True),
    Doctor(7, "Surgeon 7", True),
    Doctor(8, "Surgeon 8", True),
    Doctor(9, "Surgeon 9", False),
    Doctor(10, "Surgeon 10", False),
    Doctor(11, "Surgeon 11", False),
    Doctor(12, "Surgeon 12", False),
    Doctor(13, "Surgeon 13", True),
    Doctor(14, "Surgeon 14", True),
    Doctor(15, "Surgeon 15", True),
    Doctor(16, "Surgeon 16", False),
    Doctor(17, "Surgeon 17", True),
    Doctor(18, "Surgeon 18", True),
    Doctor(19, "Surgeon 19", True),
    Doctor(20, "Surgeon 20", False),
]

class ScheduleGenerator:
    def __init__(self, num_weeks: int = 8):
        self.num_weeks = num_weeks
        self.doctors = DOCTORS
        self.schedule = {}  # {week: {shift: [doctor_ids]}}
        
    def generate_schedule(self, max_attempts: int = 1000) -> bool:
        """Generate schedule for all weeks"""
        for week in range(self.num_weeks):
            print(f"\nGenerating schedule for Week {week + 1}...")
            success = self._generate_week_schedule(week, max_attempts)
            if not success:
                print(f"Failed to generate schedule for Week {week + 1}")
                return False
        return True
    
    def _generate_week_schedule(self, week: int, max_attempts: int) -> bool:
        """Generate schedule for a single week using greedy approach with randomization"""
        for attempt in range(max_attempts):
            week_schedule = {shift: [] for shift in range(7)}
            doctor_shifts = {doc.id: [] for doc in self.doctors}
            
            # Try to assign doctors to shifts
            success = True
            for shift_idx in range(7):
                # Try to fill this shift with available doctors
                if not self._fill_shift(shift_idx, week_schedule, doctor_shifts):
                    success = False
                    break
            
            # Check if all doctors have exactly 2 shifts
            if success and all(len(shifts) == SHIFTS_PER_WEEK 
                             for shifts in doctor_shifts.values()):
                self.schedule[week] = week_schedule
                if attempt > 0:
                    print(f"  Successfully generated after {attempt + 1} attempts")
                return True
        
        return False
    
    def _fill_shift(self, shift_idx: int, week_schedule: Dict, 
                   doctor_shifts: Dict) -> bool:
        """Fill a single shift with available doctors"""
        # Get available doctors for this shift
        available_doctors = self._get_available_doctors_for_shift(
            shift_idx, week_schedule, doctor_shifts
        )
        
        # Filter primary surgeons
        primary_available = [d for d in available_doctors 
                           if self.doctors[d-1].can_be_main_surgeon]
        
        if len(primary_available) < MIN_PRIMARY_SURGEONS_PER_SHIFT:
            return False
        
        # Randomly select doctors for this shift
        # Ensure we get at least MIN_PRIMARY_SURGEONS_PER_SHIFT primary surgeons
        random.shuffle(primary_available)
        random.shuffle(available_doctors)
        
        selected = []
        
        # First, add required primary surgeons
        selected.extend(primary_available[:MIN_PRIMARY_SURGEONS_PER_SHIFT])
        
        # Then, add 3-5 more doctors from remaining available
        # With 20 doctors × 2 shifts = 40 doctor-shifts / 7 days ≈ 5-6 doctors per shift
        remaining = [d for d in available_doctors if d not in selected]
        if remaining:
            # Need to add 3-5 more, but can't add more than available
            min_additional = min(3, len(remaining))
            max_additional = min(5, len(remaining))
            num_additional = random.randint(min_additional, max_additional)
            selected.extend(remaining[:num_additional])
        
        # Assign these doctors to the shift
        week_schedule[shift_idx] = selected
        for doc_id in selected:
            doctor_shifts[doc_id].append(shift_idx)
        
        return True
    
    def _get_available_doctors_for_shift(self, shift_idx: int, 
                                        week_schedule: Dict,
                                        doctor_shifts: Dict) -> List[int]:
        """Get list of doctors available for a given shift"""
        available = []
        
        for doc in self.doctors:
            # Check if doctor already has 2 shifts
            if len(doctor_shifts[doc.id]) >= SHIFTS_PER_WEEK:
                continue
            
            # Check consecutive days constraint (no working 2 days in a row)
            if self._violates_consecutive_days_constraint(shift_idx, doctor_shifts[doc.id]):
                continue
            
            # Check alternating days constraint
            if self._violates_alternating_constraint(shift_idx, doctor_shifts[doc.id]):
                continue
            
            # Check day-off compensation constraint (within week)
            if self._violates_day_off_constraint(shift_idx, doctor_shifts[doc.id]):
                continue
            
            # Check cross-week constraint: Monday after Sunday full-day on-call
            if self._violates_cross_week_constraint(shift_idx, doc.id):
                continue
            
            available.append(doc.id)
        
        return available
    
    def _violates_alternating_constraint(self, new_shift: int, 
                                        existing_shifts: List[int]) -> bool:
        """Check if adding new_shift would create alternating pattern (e.g., Mon + Wed)"""
        # Evening shifts are 0-4 (Mon-Fri evenings)
        if new_shift < 5:
            for existing in existing_shifts:
                if existing < 5:  # Both are evening shifts
                    # Check if they are exactly 2 days apart (alternating)
                    if abs(new_shift - existing) == 2:
                        return True
        return False
    
    def _violates_consecutive_days_constraint(self, new_shift: int, 
                                             existing_shifts: List[int]) -> bool:
        """Check if adding new_shift would create consecutive working days"""
        # Shift indices: 0=Mon evening, 1=Tue evening, ..., 4=Fri evening, 5=Sat full, 6=Sun full
        # Consecutive means shift indices differ by exactly 1
        for existing in existing_shifts:
            if abs(new_shift - existing) == 1:
                return True  # Working consecutive days is not allowed
        return False
    
    def _violates_day_off_constraint(self, new_shift: int, 
                                     existing_shifts: List[int]) -> bool:
        """Check if day-off compensation is violated"""
        # Evening shifts require next day off
        # So we need to check if doctor would need to work the next day
        
        # If new_shift is an evening (0-4), next day would be blocked
        # If new_shift is right after an evening shift, it's blocked
        
        for existing in existing_shifts:
            # If existing is an evening shift (Mon-Fri evening: 0-4)
            if existing < 5:
                next_day = existing + 1
                # Check if new_shift conflicts with the required day off
                if new_shift == next_day:
                    return True
        
        # If new_shift is an evening, check if previous shifts block it
        if new_shift < 5:
            # Check if we have a shift on the same day that would be affected
            # Actually, evening shift means day off NEXT day, so this is ok
            pass
        
        # If new_shift is Sat (5) or Sun (6), check if Fri evening is taken
        if new_shift == 5:  # Saturday
            if 4 in existing_shifts:  # Friday evening
                return True
        
        return False
    
    def _violates_cross_week_constraint(self, new_shift: int, doc_id: int) -> bool:
        """Check if Monday assignment violates Sunday full-day on-call from previous week"""
        # Only applies to Monday (shift_idx = 0) 
        if new_shift != 0:
            return False
        
        # Get current week number by checking how many weeks have been scheduled
        current_week = len(self.schedule)
        
        # Check if this is week 0 (no previous week to check)
        if current_week == 0:
            return False
        
        # Check if doctor had Sunday full-day on-call in previous week
        prev_week = current_week - 1
        prev_week_schedule = self.schedule.get(prev_week)
        
        if prev_week_schedule is None:
            return False
        
        # Check if doctor was on Sunday full-day shift (shift_idx = 6) in previous week
        sunday_doctors = prev_week_schedule.get(6, [])
        if doc_id in sunday_doctors:
            return True  # Doctor needs Monday off
        
        return False
    
    
    def _get_combinations(self, items: List, r: int) -> List[Tuple]:
        """Generate combinations of r items from items list"""
        from itertools import combinations
        return list(combinations(items, r))
    
    def validate_schedule(self) -> Tuple[bool, List[str]]:
        """Validate the entire schedule against all constraints"""
        errors = []
        
        for week in range(self.num_weeks):
            week_schedule = self.schedule[week]
            
            # Build doctor shift assignments for this week
            doctor_shifts = {doc.id: [] for doc in self.doctors}
            for shift_idx, doctors in week_schedule.items():
                for doc_id in doctors:
                    doctor_shifts[doc_id].append(shift_idx)
            
            # Check 1: Each doctor has exactly 2 shifts per week
            for doc_id, shifts in doctor_shifts.items():
                if len(shifts) != SHIFTS_PER_WEEK:
                    errors.append(
                        f"Week {week+1}: Doctor {doc_id} has {len(shifts)} shifts "
                        f"(expected {SHIFTS_PER_WEEK})"
                    )
            
            # Check 2: No alternating days
            for doc_id, shifts in doctor_shifts.items():
                ev_shifts = [s for s in shifts if s < 5]
                for i, s1 in enumerate(ev_shifts):
                    for s2 in ev_shifts[i+1:]:
                        if abs(s1 - s2) == 2:
                            errors.append(
                                f"Week {week+1}: Doctor {doc_id} has alternating "
                                f"evening shifts on days {s1} and {s2}"
                            )
            
            # Check 2.5: No consecutive working days
            for doc_id, shifts in doctor_shifts.items():
                for i, s1 in enumerate(shifts):
                    for s2 in shifts[i+1:]:
                        if abs(s1 - s2) == 1:
                            day1 = DAYS[s1]
                            day2 = DAYS[s2]
                            errors.append(
                                f"Week {week+1}: Doctor {doc_id} works consecutive days "
                                f"{day1} and {day2} (shifts {s1} and {s2})"
                            )
            
            
            # Check 3: Day-off compensation
            for doc_id, shifts in doctor_shifts.items():
                for shift in shifts:
                    if shift < 5:  # Evening shift
                        next_day = shift + 1
                        if next_day in shifts:
                            errors.append(
                                f"Week {week+1}: Doctor {doc_id} on-call evening {shift} "
                                f"and next day {next_day} (no day off)"
                            )
            
            # Check 4: Minimum primary surgeons per shift
            for shift_idx, doctors in week_schedule.items():
                primary_count = sum(1 for d in doctors 
                                  if self.doctors[d-1].can_be_main_surgeon)
                if primary_count < MIN_PRIMARY_SURGEONS_PER_SHIFT:
                    errors.append(
                        f"Week {week+1}, Shift {shift_idx}: Only {primary_count} "
                        f"primary surgeons (need {MIN_PRIMARY_SURGEONS_PER_SHIFT})"
                    )
            
            # Check 5: Cross-week constraint - Sunday full-day requires Monday off
            if week > 0:
                prev_week_schedule = self.schedule.get(week - 1)
                if prev_week_schedule:
                    sunday_doctors = prev_week_schedule.get(6, [])
                    monday_doctors = week_schedule.get(0, [])
                    
                    for doc_id in sunday_doctors:
                        if doc_id in monday_doctors:
                            errors.append(
                                f"Week {week+1}: Doctor {doc_id} had Sunday full-day "
                                f"on-call in Week {week} but is assigned to Monday "
                                f"(needs compensatory day off)"
                            )
        
        
        return len(errors) == 0, errors
    
    def export_to_excel(self, filename: str = "doctor_schedule_20doctors_8weeks.xlsx"):
        """Export schedule to Excel with formatting"""
        wb = Workbook()
        
        # Create a sheet for each week
        for week in range(self.num_weeks):
            if week == 0:
                ws = wb.active
                ws.title = f"Week {week + 1}"
            else:
                ws = wb.create_sheet(f"Week {week + 1}")
            
            self._write_week_to_sheet(ws, week)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        self._write_summary_sheet(summary_ws)
        
        wb.save(filename)
        print(f"\n[OK] Schedule exported to: {filename}")
    
    def _write_week_to_sheet(self, ws, week: int):
        """Write a single week's schedule to a worksheet"""
        week_schedule = self.schedule[week]
        
        # Define colors
        color_regular = PatternFill(start_color="E8F4F8", end_color="E8F4F8", 
                                   fill_type="solid")  # Light blue
        color_oncall = PatternFill(start_color="FFD966", end_color="FFD966", 
                                  fill_type="solid")  # Yellow
        color_dayoff = PatternFill(start_color="C6E0B4", end_color="C6E0B4", 
                                  fill_type="solid")  # Light green
        color_header = PatternFill(start_color="4472C4", end_color="4472C4", 
                                  fill_type="solid")  # Blue
        
        font_header = Font(bold=True, color="FFFFFF", size=11)
        font_bold = Font(bold=True)
        alignment_center = Alignment(horizontal="center", vertical="center", 
                                    wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Doctor'] + DAYS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = color_header
            cell.font = font_header
            cell.alignment = alignment_center
            cell.border = thin_border
        
        # Build day-off map
        dayoff_map = self._get_dayoff_map(week)
        
        # Write doctor schedules
        for doc_idx, doc in enumerate(self.doctors, 2):
            # Doctor name
            cell = ws.cell(row=doc_idx, column=1, value=doc.name)
            cell.font = font_bold
            cell.alignment = alignment_center
            cell.border = thin_border
            
            # Days
            for day_idx, day in enumerate(DAYS):
                col = day_idx + 2
                cell = ws.cell(row=doc_idx, column=col)
                
                # Determine shift type for this day
                shift_idx = self._get_shift_idx_for_day(day)
                
                if doc.id in dayoff_map and day in dayoff_map[doc.id]:
                    # Day off (compensation)
                    cell.value = "Day Off"
                    cell.fill = color_dayoff
                elif shift_idx is not None and doc.id in week_schedule.get(shift_idx, []):
                    # On-call
                    if shift_idx < 5:
                        cell.value = "On-call\n(Evening)"
                    else:
                        cell.value = "On-call\n(Full Day)"
                    cell.fill = color_oncall
                else:
                    # Regular hours (weekdays only)
                    if day in WEEKDAYS:
                        cell.value = "Regular\nHours"
                        cell.fill = color_regular
                    else:
                        cell.value = "-"
                
                cell.alignment = alignment_center
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 9):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14
        
        # Add legend
        legend_row = len(self.doctors) + 3
        ws.cell(row=legend_row, column=1, value="Legend:").font = font_bold
        
        ws.cell(row=legend_row + 1, column=1, value="Regular Hours")
        ws.cell(row=legend_row + 1, column=1).fill = color_regular
        
        ws.cell(row=legend_row + 2, column=1, value="On-call")
        ws.cell(row=legend_row + 2, column=1).fill = color_oncall
        
        ws.cell(row=legend_row + 3, column=1, value="Day Off")
        ws.cell(row=legend_row + 3, column=1).fill = color_dayoff
    
    def _get_shift_idx_for_day(self, day: str) -> int:
        """Get shift index for a given day"""
        if day == 'Monday':
            return 0
        elif day == 'Tuesday':
            return 1
        elif day == 'Wednesday':
            return 2
        elif day == 'Thursday':
            return 3
        elif day == 'Friday':
            return 4
        elif day == 'Saturday':
            return 5
        elif day == 'Sunday':
            return 6
        return None
    
    def _get_dayoff_map(self, week: int) -> Dict[int, List[str]]:
        """Get mapping of doctors to their day-off days"""
        week_schedule = self.schedule[week]
        dayoff_map = {}
        
        # Check for evening shifts in current week
        for shift_idx, doctors in week_schedule.items():
            if shift_idx < 5:  # Evening shift
                next_day_idx = shift_idx + 1
                next_day = DAYS[next_day_idx]
                
                for doc_id in doctors:
                    if doc_id not in dayoff_map:
                        dayoff_map[doc_id] = []
                    dayoff_map[doc_id].append(next_day)
        
        # Check for Sunday full-day shift from PREVIOUS week
        # Those doctors should have Monday off in THIS week
        if week > 0:
            prev_week_schedule = self.schedule.get(week - 1)
            if prev_week_schedule:
                sunday_doctors = prev_week_schedule.get(6, [])  # Sunday full day
                for doc_id in sunday_doctors:
                    if doc_id not in dayoff_map:
                        dayoff_map[doc_id] = []
                    if 'Monday' not in dayoff_map[doc_id]:
                        dayoff_map[doc_id].append('Monday')
        
        return dayoff_map
    
    def _write_summary_sheet(self, ws):
        """Write summary statistics"""
        font_header = Font(bold=True, size=12)
        font_bold = Font(bold=True)
        
        ws.cell(row=1, column=1, value="Schedule Summary").font = font_header
        
        row = 3
        ws.cell(row=row, column=1, value=f"Number of weeks: {self.num_weeks}")
        row += 1
        ws.cell(row=row, column=1, value=f"Number of doctors: {NUM_DOCTORS}")
        row += 1
        ws.cell(row=row, column=1, value=f"Shifts per week per doctor: {SHIFTS_PER_WEEK}")
        row += 2
        
        ws.cell(row=row, column=1, value="Constraints:").font = font_bold
        row += 1
        ws.cell(row=row, column=1, 
               value="• No consecutive working days")
        row += 1
        ws.cell(row=row, column=1, 
               value="• Day off after evening on-call shift")
        row += 1
        ws.cell(row=row, column=1, 
               value="• Monday off after Sunday full-day on-call shift")
        row += 1
        ws.cell(row=row, column=1, 
               value="• No alternating evening shifts (e.g., Mon + Wed)")
        row += 1
        ws.cell(row=row, column=1, 
               value=f"• At least {MIN_PRIMARY_SURGEONS_PER_SHIFT} primary surgeons "
                     f"(Surgeons 1-8, 13-15, 17-19) per shift")
        row += 2
        
        # Validation results
        is_valid, errors = self.validate_schedule()
        ws.cell(row=row, column=1, value="Validation:").font = font_bold
        row += 1
        
        if is_valid:
            ws.cell(row=row, column=1, value="✓ All constraints satisfied")
            ws.cell(row=row, column=1).font = Font(color="00B050", bold=True)
        else:
            ws.cell(row=row, column=1, value="✗ Constraint violations found:")
            ws.cell(row=row, column=1).font = Font(color="FF0000", bold=True)
            row += 1
            for error in errors[:10]:  # Show first 10 errors
                ws.cell(row=row, column=1, value=f"  • {error}")
                row += 1
        
        # Adjust column width
        ws.column_dimensions['A'].width = 60

def main():
    """Main execution function"""
    print("=" * 60)
    print("Doctor Schedule Generator")
    print("=" * 60)
    
    # Set random seed for reproducibility (optional)
    # random.seed(42)
    
    # Generate schedule for 8 weeks
    generator = ScheduleGenerator(num_weeks=8)
    
    print("\nGenerating schedules...")
    success = generator.generate_schedule(max_attempts=1000)
    
    if not success:
        print("\n[ERROR] Failed to generate valid schedule")
        return
    
    print("\n[OK] Schedule generation complete")
    
    # Validate
    print("\nValidating schedule...")
    is_valid, errors = generator.validate_schedule()
    
    if is_valid:
        print("[OK] All constraints satisfied")
    else:
        print(f"[ERROR] Found {len(errors)} constraint violations:")
        for error in errors[:5]:
            print(f"  • {error}")
        if len(errors) > 5:
            print(f"  ... and {len(errors) - 5} more")
    
    # Export to Excel
    print("\nExporting to Excel...")
    generator.export_to_excel("doctor_schedule_20doctors_8weeks.xlsx")
    
    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)

if __name__ == "__main__":
    main()
